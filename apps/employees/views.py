from rest_framework import viewsets
from .models import Department, Designation, Category, EmployeeType, EmployeeProfile, EmployeeAllowance, EmployeeDeduction
from .serializers import *
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework import serializers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
import io

class AdminTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        data = super().validate(attrs)
        if not self.user.is_staff:
            raise serializers.ValidationError("You are not authorized as an admin.")
        return data

class AdminLoginView(TokenObtainPairView):
    serializer_class = AdminTokenObtainPairSerializer


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

class DesignationViewSet(viewsets.ModelViewSet):
    queryset = Designation.objects.all()
    serializer_class = DesignationSerializer

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

class EmployeeTypeViewSet(viewsets.ModelViewSet):
    queryset = EmployeeType.objects.all()
    serializer_class = EmployeeTypeSerializer

class EmployeeProfileViewSet(viewsets.ModelViewSet):
    queryset = EmployeeProfile.objects.all()
    serializer_class = EmployeeProfileSerializer

class EmployeeAllowanceViewSet(viewsets.ModelViewSet):
    queryset = EmployeeAllowance.objects.all()
    serializer_class = EmployeeAllowanceSerializer

class EmployeeDeductionViewSet(viewsets.ModelViewSet):
    queryset = EmployeeDeduction.objects.all()
    serializer_class = EmployeeDeductionSerializer

class EmployeesReportAPIView(APIView):
    def get(self, request):
        data = []
        qs = EmployeeProfile.objects.select_related('department', 'designation').all()
        for e in qs:
            data.append({
                "id": e.id,
                "name": e.name,
                "employee_code": e.employee_code,
                "department": e.department.name if e.department else None,
                "designation": e.designation.title if e.designation else None,
                "status": e.status,
                "net_salary": str(e.net_salary),
                "date_of_joining": e.date_of_joining,
            })
        return Response(data)


class EmployeesReportPDFAPIView(APIView):
    def get(self, request):
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return HttpResponse("PDF generation library not installed", status=501)
        qs = EmployeeProfile.objects.select_related('department', 'designation').all()
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph("Employees list", styles["Heading1"]))
        elements.append(Spacer(1, 12))
        data = [["ID", "Name", "Code", "Department", "Designation", "Status", "Net Salary", "DOJ"]]
        for e in qs:
            data.append([
                e.id,
                e.name or "",
                e.employee_code or "",
                e.department.name if e.department else "",
                e.designation.title if e.designation else "",
                e.status,
                str(e.net_salary),
                e.date_of_joining.isoformat() if e.date_of_joining else "",
            ])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ]))
        elements.append(table)
        doc.build(elements)
        from datetime import datetime
        year = datetime.now().year
        resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        resp['Content-Disposition'] = f'attachment; filename="employees_list_{year}.pdf"'
        return resp

class EmployeesReportExcelAPIView(APIView):
    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
        except ImportError:
            return HttpResponse("Excel generation library not installed", status=501)
        qs = EmployeeProfile.objects.select_related('department', 'designation').all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        headers = ["ID", "Name", "Code", "Department", "Designation", "Status", "Net Salary", "DOJ"]
        transformed = []
        for h in headers:
            parts = h.split()
            if len(parts) == 2:
                transformed.append(f"{parts[0]}\n{parts[1]}")
            else:
                transformed.append(h)
        ws.append(transformed)
        for e in qs:
            ws.append([
                e.id,
                e.name or "",
                e.employee_code or "",
                e.department.name if e.department else "",
                e.designation.title if e.designation else "",
                e.status,
                float(e.net_salary or 0),
                (e.date_of_joining.isoformat() if e.date_of_joining else ""),
            ])
        for cell in ws[1]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[column].width = min(max_length + 2, 25)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp['Content-Disposition'] = 'attachment; filename="employees_list.xlsx"'
        return resp

class EmployeeDeductionsReportAPIView(APIView):
    def get(self, request):
        year = request.query_params.get("year")
        month = request.query_params.get("month")
        qs = EmployeeDeduction.objects.select_related('employee', 'deduction_type').all()
        if year and month:
            qs = qs.filter(date__year=int(year), date__month=int(month))
        data = []
        for d in qs:
            data.append({
                "id": d.id,
                "employee_id": d.employee.id,
                "employee_name": d.employee.name,
                "employee_code": d.employee.employee_code,
                "deduction_type": d.deduction_type.name if d.deduction_type else None,
                "amount": str(d.amount),
                "method": d.method,
                "months": d.months,
                "date": d.date,
                "reimbursed_amount": str(d.reimbursed_amount or 0),
                "remaining_amount": str(d.remaining_amount or 0),
                "is_closed": d.is_closed,
            })
        return Response(data)


class EmployeeDeductionsReportPDFAPIView(APIView):
    def get(self, request):
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            import calendar
        except ImportError:
            return HttpResponse("PDF generation library not installed", status=501)
        year = request.query_params.get("year")
        month = request.query_params.get("month")
        qs = EmployeeDeduction.objects.select_related('employee', 'deduction_type').all()
        if year and month:
            qs = qs.filter(date__year=int(year), date__month=int(month))
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        if year and month:
            mname = calendar.month_name[int(month)]
            elements.append(Paragraph(f"Employee Deductions report of month {mname} {year}", styles["Heading1"]))
        else:
            elements.append(Paragraph("Employee Deductions report", styles["Heading1"]))
        elements.append(Spacer(1, 12))
        data = [["ID", "Employee", "Code", "Type", "Amount", "Method", "Months", "Date", "Reimbursed", "Remaining", "Closed"]]
        for d in qs:
            data.append([
                d.id,
                d.employee.name,
                d.employee.employee_code,
                d.deduction_type.name if d.deduction_type else "",
                str(d.amount),
                d.method,
                d.months or "",
                d.date.isoformat() if d.date else "",
                str(d.reimbursed_amount or 0),
                str(d.remaining_amount or 0),
                d.is_closed,
            ])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ]))
        elements.append(table)
        doc.build(elements)
        resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        if year and month:
            resp['Content-Disposition'] = f'attachment; filename="employee_deductions_report_{int(month):02d}_{year}.pdf"'
        else:
            resp['Content-Disposition'] = 'attachment; filename="employee_deductions_report.pdf"'
        return resp

class EmployeeDeductionsReportExcelAPIView(APIView):
    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
        except ImportError:
            return HttpResponse("Excel generation library not installed", status=501)
        year = request.query_params.get("year")
        month = request.query_params.get("month")
        qs = EmployeeDeduction.objects.select_related('employee', 'deduction_type').all()
        if year and month:
            qs = qs.filter(date__year=int(year), date__month=int(month))
        wb = Workbook()
        ws = wb.active
        title = "Employee Deductions"
        if year and month:
            title = f"Deductions {int(month):02d}-{year}"
        ws.title = title[:31]
        headers = ["ID", "Employee", "Code", "Type", "Amount", "Method", "Months", "Date", "Reimbursed", "Remaining", "Closed"]
        transformed = []
        for h in headers:
            parts = h.split()
            if len(parts) == 2:
                transformed.append(f"{parts[0]}\n{parts[1]}")
            else:
                transformed.append(h)
        ws.append(transformed)
        for d in qs:
            ws.append([
                d.id,
                d.employee.name,
                d.employee.employee_code,
                d.deduction_type.name if d.deduction_type else "",
                float(d.amount or 0),
                d.method,
                d.months or "",
                (d.date.isoformat() if d.date else ""),
                float(d.reimbursed_amount or 0),
                float(d.remaining_amount or 0),
                d.is_closed,
            ])
        for cell in ws[1]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[column].width = min(max_length + 2, 25)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if year and month:
            resp['Content-Disposition'] = f'attachment; filename="employee_deductions_report_{int(month):02d}_{year}.xlsx"'
        else:
            resp['Content-Disposition'] = 'attachment; filename="employee_deductions_report.xlsx"'
        return resp
