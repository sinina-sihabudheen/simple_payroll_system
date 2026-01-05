from rest_framework import viewsets
from rest_framework.views import APIView
from .models import Attendance, Leave, LeaveType, EsslPunch, EsslConfig
from apps.employees.models import EmployeeProfile  
from rest_framework.response import Response
from datetime import date, datetime
from rest_framework import status
from .serializers import *
from collections import defaultdict
from django.http import HttpResponse
import io

class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer

class LeaveViewSet(viewsets.ModelViewSet):
    queryset = Leave.objects.all()
    serializer_class = LeaveSerializer
    
    def perform_update(self, serializer):
        leave = serializer.save()
        if leave.approved:
            Attendance.objects.update_or_create(
                employee=leave.employee,
                date=leave.date,
                defaults={
                    "is_present": True,
                    "marked_manually": False,
                    "in_time": None,
                    "out_time": None,
                }
            )

class LeaveTypeViewSet(viewsets.ModelViewSet):
    queryset = LeaveType.objects.all()
    serializer_class = LeaveTypeSerializer

class EsslPunchViewSet(viewsets.ModelViewSet):
    queryset = EsslPunch.objects.all()
    serializer_class = EsslPunchSerializer

from .utils.essl_reader import fetch_essl_data

class SyncEsslToAttendance(APIView):
    def post(self, request):
        # First, try to fetch new data from the device
        success, message = fetch_essl_data()
        
        # Even if fetching fails (e.g., device offline), we process existing logs
        # but we should warn the user.
        
        punch_map = defaultdict(list)
        for punch in EsslPunch.objects.all():
            punch_date = punch.punch_time.date()
            punch_map[(punch.employee_code, punch_date)].append(punch.punch_time.time())

        updated_count = 0
        for (emp_code, punch_date), times in punch_map.items():
            try:
                employee = EmployeeProfile.objects.get(employee_code=emp_code)
                Attendance.objects.update_or_create(
                    employee=employee,
                    date=punch_date,
                    defaults={
                        'in_time': min(times),
                        'out_time': max(times),
                        'is_present': True,
                        'marked_manually': False
                    }
                )
                updated_count += 1
            except EmployeeProfile.DoesNotExist:
                continue

        response_data = {
            'detail': 'ESSL synced to attendance.',
            'device_status': message,
            'records_processed': updated_count
        }
        
        status_code = status.HTTP_200_OK if success else status.HTTP_207_MULTI_STATUS
        return Response(response_data, status=status_code)

class EsslConfigView(APIView):
    def get(self, request):
        cfg = EsslConfig.objects.first()
        if not cfg:
            cfg = EsslConfig.objects.create()
        serializer = EsslConfigSerializer(cfg)
        return Response(serializer.data)

    def post(self, request):
        cfg = EsslConfig.objects.first()
        if not cfg:
            cfg = EsslConfig()
        serializer = EsslConfigSerializer(cfg, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class AttendanceByDate(APIView):
    def get(self, request):
        date_str = request.query_params.get("date")
        if not date_str:
            return Response({"error": "Date parameter is required."}, status=400)

        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        employees = EmployeeProfile.objects.filter(date_of_joining__lte=date_obj)

        response_data = []
        for emp in employees:
            attendance = Attendance.objects.filter(employee=emp, date=date_obj).first()
            response_data.append({
                "employee_id": emp.id,
                "employee_name": f"{emp.name}",
                "employee_code": emp.employee_code,
                "attendance": AttendanceSerializer(attendance).data if attendance else None
            })

        return Response(response_data)

class MarkAttendanceManually(APIView):
    def post(self, request):
        data = request.data

        employee_id = data.get("employee")
        attendance_date = data.get("date")
        is_present = data.get("is_present", False)
        in_time = data.get("in_time")
        out_time = data.get("out_time")
        marked_manually = data.get("marked_manually", True)

        if not employee_id or not attendance_date:
            return Response({"error": "Employee and date are required."}, status=400)

        try:
            employee = EmployeeProfile.objects.get(id=employee_id)
        except EmployeeProfile.DoesNotExist:
            return Response({"error": "Employee not found."}, status=404)

        try:
            attendance_date_obj = datetime.strptime(attendance_date, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        if attendance_date_obj > date.today():
            return Response({"error": "Cannot mark attendance for a future date."}, status=400)

        if attendance_date_obj < employee.date_of_joining:
            return Response({"error": "Cannot mark attendance before employee's date of joining."}, status=400)

        if is_present and not in_time:
            return Response({"error": "In time is required for attendance."}, status=400)

        attendance, created = Attendance.objects.update_or_create(
            employee=employee,
            date=attendance_date,
            defaults={
                "in_time": in_time,
                "out_time": out_time,
                "is_present": is_present,
                "marked_manually": marked_manually,
            },
        )

        serializer = AttendanceSerializer(attendance)
        return Response(serializer.data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

class AttendanceMonthlyReportAPIView(APIView):
    def get(self, request):
        year = int(request.query_params.get("year"))
        month = int(request.query_params.get("month"))
        employees = EmployeeProfile.objects.all()
        data = []
        for emp in employees:
            qs = Attendance.objects.filter(employee=emp, date__year=year, date__month=month)
            present_days = qs.filter(is_present=True).count()
            total_days = qs.count()
            absent_days = max(0, total_days - present_days)
            data.append({
                "employee_id": emp.id,
                "employee_name": emp.name,
                "employee_code": emp.employee_code,
                "present_days": present_days,
                "absent_days": absent_days,
                "month": month,
                "year": year,
            })
        return Response(data)


class AttendanceMonthlyReportPDFAPIView(APIView):
    def get(self, request):
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            import calendar
        except ImportError:
            return HttpResponse("PDF generation library not installed", status=501)
        year = int(request.query_params.get("year"))
        month = int(request.query_params.get("month"))
        employees = EmployeeProfile.objects.all()
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        mname = calendar.month_name[month]
        elements.append(Paragraph(f"Attendance report of month {mname} {year}", styles["Heading1"]))
        elements.append(Spacer(1, 12))
        data = [["Employee", "Code", "Present Days", "Absent Days"]]
        for emp in employees:
            qs = Attendance.objects.filter(employee=emp, date__year=year, date__month=month)
            present_days = qs.filter(is_present=True).count()
            total_days = qs.count()
            absent_days = max(0, total_days - present_days)
            data.append([emp.name, emp.employee_code, present_days, absent_days])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ]))
        elements.append(table)
        doc.build(elements)
        resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        resp['Content-Disposition'] = 'attachment; filename="attendance.pdf"'
        resp['Content-Disposition'] = f'attachment; filename="attendance_report_{month:02d}_{year}.pdf"'
        return resp

class AttendanceMonthlyReportExcelAPIView(APIView):
    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
            import calendar
        except ImportError:
            return HttpResponse("Excel generation library not installed", status=501)
        year = int(request.query_params.get("year"))
        month = int(request.query_params.get("month"))
        employees = EmployeeProfile.objects.all()
        wb = Workbook()
        ws = wb.active
        mname = calendar.month_name[month]
        ws.title = f"Attendance {month:02d}-{year}"[:31]
        headers = ["Employee", "Code", "Present Days", "Absent Days"]
        transformed = []
        for h in headers:
            parts = h.split()
            if len(parts) == 2:
                transformed.append(f"{parts[0]}\n{parts[1]}")
            else:
                transformed.append(h)
        ws.append(transformed)
        for emp in employees:
            qs = Attendance.objects.filter(employee=emp, date__year=year, date__month=month)
            present_days = qs.filter(is_present=True).count()
            total_days = qs.count()
            absent_days = max(0, total_days - present_days)
            ws.append([emp.name, emp.employee_code, present_days, absent_days])
        for cell in ws[1]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[column].width = min(max_length + 2, 25)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp['Content-Disposition'] = f'attachment; filename="attendance_report_{month:02d}_{year}.xlsx"'
        return resp
