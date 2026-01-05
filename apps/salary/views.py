from rest_framework.views import APIView
from rest_framework.response import Response
from apps.attendance.models import Attendance, Leave
from apps.employees.models import EmployeeProfile
from .models import SalaryRecord, Allowance, Deduction
from .serializers import AllowanceSerializer, SalaryRecordSerializer, DeductionSerializer
from .utils import calculate_deductions, calculate_working_days
from datetime import date
from calendar import monthrange
from rest_framework import viewsets
from decimal import Decimal
from apps.employees.models import EmployeeAllowance, EmployeeDeduction
from django.db.models import Sum
from django.utils import timezone
from rest_framework import status
from django.http import HttpResponse
import io



class AllowanceViewSet(viewsets.ModelViewSet):
    queryset = Allowance.objects.all()
    serializer_class = AllowanceSerializer

class DeductionViewSet(viewsets.ModelViewSet):
    queryset = Deduction.objects.all()
    serializer_class = DeductionSerializer


def compute_salary_results(year, month):
    today = date.today()
    if year == today.year and month == today.month:
        last_day = today.day
    else:
        last_day = monthrange(year, month)[1]
    results = []
    employees = EmployeeProfile.objects.all()
    for employee in employees:
        present_days, absent_days, sunday_count, approved_leave_count = calculate_working_days(employee, year, month, last_day)
        basic = employee.basic_salary or Decimal(0)
        hra = employee.house_rent_allowance or Decimal(0)
        transport = employee.transportation_allowance or Decimal(0)
        col_allowance = employee.cost_of_living_allowance or Decimal(0)
        total_allowance = hra + transport + col_allowance
        gross_basic = basic + total_allowance
        total_days_in_month = monthrange(year, month)[1]
        daily_salary = gross_basic / total_days_in_month
        paid_days = present_days + sunday_count + approved_leave_count
        salary_on_attendance = daily_salary * paid_days
        advance_deduction, other_deductions = calculate_deductions(employee, year, month)
        total_deductions = advance_deduction + other_deductions
        total_salary = salary_on_attendance - total_deductions
        previous_due = SalaryRecord.objects.filter(employee=employee, status="pending").exclude(year=year, month=month).aggregate(total_due=Sum('gross_salary'))['total_due'] or Decimal(0)
        salary_record, created = SalaryRecord.objects.get_or_create(
            employee=employee,
            year=year,
            month=month,
            defaults={
                'present_days': present_days,
                'absent_days': absent_days,
                'lop_count': absent_days,
                'total_allowances': total_allowance,
                'total_deductions': total_deductions,
                'gross_salary': total_salary,
                'salary_due': previous_due,
                'status': 'pending',
            }
        )
        if not created:
            salary_record.present_days = present_days
            salary_record.absent_days = absent_days
            salary_record.lop_count = absent_days
            salary_record.total_allowances = total_allowance
            salary_record.total_deductions = total_deductions
            salary_record.gross_salary = total_salary
            salary_record.salary_due = previous_due
            salary_record.save()
        results.append({
            'id': salary_record.id,
            'employee_id': employee.id,
            'employee_name': employee.name,
            'employee_number': employee.employee_code,
            'present_days': present_days,
            'absent_days': absent_days,
            'leave_count': sunday_count,
            'approved_leave_count': approved_leave_count,
            'basic_salary': basic,
            'gross_basic_salary': gross_basic,
            'house_rent_allowance': hra,
            'transportation_allowance': transport,
            'cost_of_living_allowance': col_allowance,
            'total_allowance': total_allowance,
            'salary_on_attendance': salary_on_attendance,
            'advance_deduction': advance_deduction,
            'other_deductions': other_deductions,
            'total_deduction': total_deductions,
            'total_salary': total_salary,
            'holiday_count': sunday_count,
            'status': salary_record.status,
            'paid_amount': salary_record.paid_amount or Decimal(0),
            'balance_amount': salary_record.balance_amount or ((salary_record.gross_salary + salary_record.salary_due) - (salary_record.paid_amount or Decimal(0))),
            'paid_date': salary_record.paid_date,
        })
    return results

class GenerateSalaryAPIView(APIView):
    def post(self, request):
        year = int(request.data.get('year'))
        month = int(request.data.get('month'))
        results = compute_salary_results(year, month)
        return Response(results)


class PaySalaryAPIView(APIView):
    def patch(self, request, pk):
        try:
            salary_record = SalaryRecord.objects.get(pk=pk)
        except SalaryRecord.DoesNotExist:
            return Response({"error": "Salary record not found"}, status=status.HTTP_404_NOT_FOUND)

        paid_amount = Decimal(request.data.get("paid_amount", 0))
        total_to_pay = salary_record.gross_salary + salary_record.salary_due

        salary_record.paid_amount += paid_amount

        # Deduction reimbursement
        remaining_amount_to_apply = paid_amount
        deductions = EmployeeDeduction.objects.filter(employee=salary_record.employee, is_closed=False)
        for deduction in deductions:
            applied = deduction.apply_reimbursement(remaining_amount_to_apply)
            remaining_amount_to_apply -= applied
            if remaining_amount_to_apply <= 0:
                break

        if salary_record.paid_amount >= total_to_pay:
            salary_record.balance_amount = Decimal(0)
            salary_record.status = "paid"
            salary_record.paid_date = timezone.now().date()
        else:
            salary_record.balance_amount = total_to_pay - salary_record.paid_amount
            salary_record.status = "partially_paid"
            salary_record.paid_date = timezone.now().date()

        salary_record.save()

        return Response({
            "employee": salary_record.employee.name,
            "month": salary_record.month,
            "year": salary_record.year,
            "gross_salary": salary_record.gross_salary,
            "salary_due": salary_record.salary_due,
            "paid_amount": salary_record.paid_amount,
            "balance_amount": salary_record.balance_amount,
            "status": salary_record.status,
            "paid_date": salary_record.paid_date,
        }, status=status.HTTP_200_OK)


class SalaryReportPDFAPIView(APIView):
    def get(self, request):
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            import calendar
        except ImportError:
            return HttpResponse("PDF generation library not installed", status=501)
        year = int(request.query_params.get('year'))
        month = int(request.query_params.get('month'))
        compute_salary_results(year, month)
        qs = SalaryRecord.objects.select_related('employee').filter(year=year, month=month)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = []
        mname = calendar.month_name[month]
        elements.append(Paragraph(f"Salary report of month {mname} {year}", styles["Heading1"]))
        elements.append(Spacer(1, 12))
        header_labels = ["Employee", "Code", "Year", "Month", "Present", "Absent", "LOP", "Total Allowances", "Total Deductions", "Gross Salary", "Salary Due", "Paid Amount", "Balance", "Status", "Paid Date", "Generated On"]
        from reportlab.lib.styles import ParagraphStyle
        header_style = ParagraphStyle("HeaderCell", parent=styles["Normal"], alignment=1, fontName="Helvetica-Bold", fontSize=8, leading=9)
        header_cells = []
        for h in header_labels:
            parts = h.split()
            if len(parts) == 2:
                header_cells.append(Paragraph(f"{parts[0]}<br/>{parts[1]}", header_style))
            else:
                header_cells.append(Paragraph(h, header_style))
        table_data = [header_cells]
        for s in qs:
            table_data.append([
                s.employee.name,
                s.employee.employee_code,
                s.year,
                s.month,
                s.present_days,
                s.absent_days,
                s.lop_count,
                str(s.total_allowances),
                str(s.total_deductions),
                str(s.gross_salary),
                str(s.salary_due),
                str(s.paid_amount),
                str(s.balance_amount),
                s.status,
                s.paid_date.isoformat() if s.paid_date else "",
                (s.generated_on.date().isoformat() if s.generated_on else ""),
            ])
        col_widths = [80, 50, 30, 30, 40, 40, 35, 50, 50, 50, 50, 50, 50, 45, 60, 60]
        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (7, 1), (13, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ]))
        elements.append(table)
        doc.build(elements)
        resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        resp['Content-Disposition'] = f'attachment; filename="salary_report_{month:02d}_{year}.pdf"'
        return resp

class SalaryReportExcelAPIView(APIView):
    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("Excel generation library not installed", status=501)
        year = int(request.query_params.get('year'))
        month = int(request.query_params.get('month'))
        compute_salary_results(year, month)
        qs = SalaryRecord.objects.select_related('employee').filter(year=year, month=month)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Salaries {month:02d}-{year}"
        headers = ["ID","Employee","Code","Year","Month","Present Days","Absent Days","LOP","Total Allowances","Total Deductions","Gross Salary","Salary Due","Paid Amount","Balance Amount","Status","Paid Date","Generated On"]
        transformed = []
        for h in headers:
            parts = h.split()
            if len(parts) == 2:
                transformed.append(f"{parts[0]}\n{parts[1]}")
            else:
                transformed.append(h)
        ws.append(transformed)
        for s in qs:
            ws.append([
                s.id,
                s.employee.name,
                s.employee.employee_code,
                s.year,
                s.month,
                s.present_days,
                s.absent_days,
                s.lop_count,
                float(s.total_allowances or 0),
                float(s.total_deductions or 0),
                float(s.gross_salary or 0),
                float(s.salary_due or 0),
                float(s.paid_amount or 0),
                float(s.balance_amount or 0),
                s.status,
                s.paid_date.isoformat() if s.paid_date else "",
                (s.generated_on.date().isoformat() if s.generated_on else ""),
            ])
        from openpyxl.styles import Alignment
        for cell in ws[1]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[column].width = min(max_length + 2, 25)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp['Content-Disposition'] = f'attachment; filename="salary_report_{month:02d}_{year}.xlsx"'
        return resp
